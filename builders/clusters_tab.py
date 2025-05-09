from __future__ import annotations
import json
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Bring in the symbol‐priority map from hosts_tab
from builders.hosts_tab import PRIO2SYM
SYM2PRIO = {v: k for k, v in PRIO2SYM.items()}


def build_clusters_df(
    df_hosts: pd.DataFrame,
    virtual_clusters_json: str | Path,
) -> pd.DataFrame:
    """
    Summarize across clusters:
      • Sum cores & sockets of ALL hosts in each cluster (from Virtualization Clusters JSON)
      • Pull datacenter, VMotion support, and VI SDK server from raw_data
      • Roll up each option symbol (highest‐priority) across Oracle hosts in the cluster
    """
    # ── Load cluster specs from Virtualization Clusters.json ────────────────
    try:
        df_specs = pd.read_json(virtual_clusters_json)
    except (FileNotFoundError, ValueError):
        # Missing or invalid virtualization clusters JSON: return empty sheet
        id_block = [
            "visdk_server",
            "cluster_name",
            "instance_uuid",
            "number_of_hosts",
            "admission_control_enabled",
            "num_vmotions",
            "ha_enabled",
            "drs_enabled",
            "overall_status",
            "total_number_of_processors",
            "total_number_of_cores",
        ]
        # Any option columns from hosts DataFrame
        opt_block = sorted([c for c in df_hosts.columns if c not in id_block])
        return pd.DataFrame(columns=id_block + opt_block)
    # Rename device_name -> cluster_name, number_of_cluster_members -> number_of_hosts
    df_specs = df_specs.rename(columns={
        "device_name": "cluster_name",
        "number_of_cluster_members": "number_of_hosts"
    })

    # Map raw_data keys to dataframe column names
    RAW_KEY_MAP = {
        "Config status": "config_status",
        "OverallStatus": "overall_status",
        "NumHosts": "number_of_hosts",            # can override if needed
        "NumCpuCores": "total_number_of_cores",
        "NumCpuThreads": "total_number_of_threads",
        "Num Vmotions": "num_vmotions",
        "HA Enabled": "ha_enabled",
        "DRS enabled": "drs_enabled",
        "AdmissionControlEnabled": "admission_control_enabled",
        "FailoverLevel": "failover_level",
        "VI SDK Server": "visdk_server",
        "InstanceUUID": "instance_uuid",
    }

    # Extract fields defined in RAW_KEY_MAP from raw_data JSON
    def _extract_cluster_fields(r: pd.Series) -> pd.Series:
        raw = r.get("raw_data", "")
        try:
            payload = json.loads(raw) if raw and isinstance(raw, str) else (raw if isinstance(raw, dict) else {})
        except Exception:
            import ast
            try:
                payload = ast.literal_eval(raw) if raw else {}
            except Exception:
                payload = {}
        # Map each raw key to its column name
        for raw_key, col_name in RAW_KEY_MAP.items():
            # Only fill if column not already populated
            if col_name not in r or not r.get(col_name):
                r[col_name] = payload.get(raw_key, "")
        return r

    df_specs = df_specs.apply(_extract_cluster_fields, axis=1)
    # Pick only the relevant fields (if present)
    keep = [
        "cluster_name",
        "number_of_hosts",
        "total_number_of_processors",
        "total_number_of_cores",
        "total_number_of_threads",
        "overall_status",
        "config_status",
        "ha_enabled",
        "drs_enabled",
        "num_vmotions",
        "failover_level",
        "admission_control_enabled",
        "visdk_server",
        "instance_uuid",
    ]
    keep = [c for c in keep if c in df_specs.columns]
    # One row per cluster
    specs_grp = df_specs[keep].drop_duplicates(subset=["cluster_name"])

    # ── 3) Roll up option symbols from df_hosts (only Oracle hosts) ────────
    # Identify option columns in the host‐level sheet
    non_opts = {
        "cluster_name", "physical_device", "number_of_vms",
        "operating_system_type", "esx_version", "manufacturer",
        "model", "cpu_model", "total_number_of_processors",
        "cores_per_cpu", "total_number_of_cores",
    }
    opt_cols = [c for c in df_hosts.columns if c not in non_opts]

    def _cluster_opts(sub: pd.DataFrame) -> pd.Series:
        out = {}
        for c in opt_cols:
            syms = sub[c].dropna().tolist()
            if syms:
                out[c] = max(syms, key=lambda s: SYM2PRIO.get(s, 0))
            else:
                out[c] = ""
        return pd.Series(out)

    opts_grp = df_hosts.groupby("cluster_name", as_index=False).apply(_cluster_opts).reset_index(drop=True)

    # ── 4) Merge sizing+metadata with option symbols ────────────────────────
    df = specs_grp.merge(opts_grp, on="cluster_name", how="left")
    # ── Drop unwanted columns from Clusters tab ────────────────────────────
    df = df.drop(columns=[
        "config_status",
        "failover_level",
        "total_number_of_threads",
    ], errors="ignore")

    # ── 5) Reorder: identity, metadata, sizing, then options ──────────────
    id_block = [
        "visdk_server",
        "cluster_name",
        "instance_uuid",
        "number_of_hosts",
        "admission_control_enabled",
        "num_vmotions",
        "ha_enabled",
        "drs_enabled",
        "overall_status",
        "total_number_of_processors",
        "total_number_of_cores",
    ]
    id_block = [c for c in id_block if c in df.columns]
    opt_block = sorted([c for c in df.columns if c not in id_block])
    return df[id_block + opt_block].fillna("")


def write_clusters_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    sheet_name: str = "Clusters",
) -> None:
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    ws.freeze_panes = ws["A2"]


def build(
    df_hosts: pd.DataFrame,
    virtual_clusters_json: str | Path,
    workbook: Workbook | None = None,
) -> tuple[Workbook, pd.DataFrame]:
    wb = workbook or Workbook()
    df_clusters = build_clusters_df(df_hosts, virtual_clusters_json)
    write_clusters_sheet(wb, df_clusters)
    return wb, df_clusters