# builders/db_details_tab.py
"""
Build the Database Details tab
------------------------------

Inputs
~~~~~~
1. oracle_dbs_json         -> "Oracle Databases Installed.json"
   * one object per DB instance (device_name, database_name, edition, …)

2. options_evidence_json   -> "Options Usage Evidence.json"
   * one object per (device, db, option) with 'result' in
     {"used","historical","cloned","verify",""}.

Output
~~~~~~
* df_db_details  (DataFrame)  –> returned to caller
* "Database Details" sheet    –> written into workbook (optional)
"""

from pathlib import Path
from typing   import Literal, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ── 1.  Priority & symbol maps  ──────────────────────────────────────────────
OPTION_PRIORITY: Dict[str, int] = {
    "used": 4,
    "historical": 3,
    "cloned": 2,
    "verify": 1,
    "": 0,          # empty / unknown
    None: 0
}
OPTION_SYMBOL: Dict[str, str] = {
    "used": "+",
    "historical": "~",
    "cloned": "#",
    "verify": "^",
    "": "",
    None: ""
}

# ── 2.  Pure transform function  ─────────────────────────────────────────────
def build_db_details_df(
    oracle_dbs_json: Path | str,
    options_evidence_json: Path | str,
) -> pd.DataFrame:
    """
    Returns a DataFrame with:
        core columns  … device, database, edition, version, etc.
        option cols   … one col per option, holding +  ~  #  ^  or "".
    """

   # ---- 2.1  base table = installed databases -----------------------------
    df_base = pd.read_json(oracle_dbs_json)

    # normalise key columns
    df_base["device_key"] = df_base["device_name"].str.lower().str.strip()
    df_base["db_key"]     = df_base["database_name"].str.lower().str.strip()

    # ---- 2.2  option evidence  ---------------------------------------------
    df_opt = pd.read_json(options_evidence_json)
 
    # same keys
    df_opt["device_key"] = df_opt["device_name"].str.lower().str.strip()
    df_opt["db_key"]     = df_opt["database_name"].str.lower().str.strip()
    df_opt["opt_key"]    = df_opt["option_name"].str.lower().str.strip()

    # choose best result per (device, db, option)
    df_opt["prio"] = df_opt["result"].str.lower().map(OPTION_PRIORITY)
    df_best = (
        df_opt
        .sort_values("prio", ascending=False)
        .drop_duplicates(subset=["device_key", "db_key", "opt_key"])
    )

    # pivot → wide matrix of symbols
    df_best["symbol"] = df_best["result"].str.lower().map(OPTION_SYMBOL)
    df_matrix = (
        df_best
        .pivot(index=["device_key", "db_key"],
               columns="opt_key",
               values="symbol")
        .fillna("")
        .reset_index()
    )

    # ── 2.4  merge dbid from evidence JSON -------------------------------
    # extract unique dbid per (device_key, db_key)
    df_dbid = df_opt[["device_key", "db_key", "dbid"]].drop_duplicates()
    # merge it into the base DataFrame
    df_base = df_base.merge(df_dbid, on=["device_key", "db_key"], how="left")

    # ---- 2.3  merge matrix back onto base ----------------------------------
    df_final = (
        df_base
        .merge(df_matrix, on=["device_key", "db_key"], how="left")
        .drop(columns=["device_key", "db_key"])
        .fillna("")          # any missing option = ""
    )

    # expose VM name explicitly so later builders can group by it
    df_final["virtual_device"] = df_final["device_name"]

    # ── reorder columns: remove unwanted, then core, then RAC fields, then the rest ──
    # drop 'product_name' if present
    if "product_name" in df_final.columns:
        df_final = df_final.drop(columns=["product_name"])

    core_cols = [
        "device_name", "dbid", "database_name", "product_edition",
        "product_version", "full_version", "instance_status",
        "goldengate_enabled", "source"
    ]
    # RAC-related fields to move next
    rac_cols = [c for c in ["rac_hosts", "rac_instances", "rac_members_count"] if c in df_final.columns]
    # all other columns
    other_cols = [c for c in df_final.columns if c not in core_cols + rac_cols]
    df_final = df_final[core_cols + rac_cols + other_cols]

    # ── remove duplicate DB rows per device+name, keep one with dbid if present ──
    if "dbid" in df_final.columns:
        # mark rows with dbid
        df_final["_keep"] = df_final["dbid"].astype(bool)
        # sort so rows with dbid come first
        df_final = df_final.sort_values("_keep", ascending=False)
        # drop duplicates, keeping the first (which will have dbid if any)
        df_final = df_final.drop_duplicates(subset=["device_name", "database_name"], keep="first")
        # clean up helper column
        df_final = df_final.drop(columns=["_keep"])

    return df_final


# ── 3.  Writer helper  ───────────────────────────────────────────────────────
def write_db_details_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    sheet_name: str = "Database Details"
) -> None:
    """
    Writes df (without virtual_device) to wb[sheet_name]. If the sheet exists it is replaced.
    """

    # drop helper column before writing
    df_to_write = df.copy()
    if "virtual_device" in df_to_write.columns:
        df_to_write.drop(columns=["virtual_device"], inplace=True)

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)

    for row in dataframe_to_rows(df_to_write, index=False, header=True):
        ws.append(row)

    ws.freeze_panes = ws["A2"]


# ── 4.  Convenience `build()` front‑door  ────────────────────────────────────
def build(
    oracle_dbs_json: Path | str,
    options_evidence_json: Path | str,
    workbook: Workbook | None = None
) -> tuple[Workbook, pd.DataFrame]:
    """
    Orchestrates transform + write.
    Returns (workbook, df_db_details)
    """
    df = build_db_details_df(oracle_dbs_json, options_evidence_json)

    wb = workbook or Workbook()
    write_db_details_sheet(wb, df)

    return wb, df