# builders/virtual_devices_tab.py
#
# One‑row‑per‑virtual‑machine tab, built directly from *Virtual Devices.json*
# (no dependence on All Devices.json).  Mirrors the simpler logic of
# hosts_tab: load → normalise → unpack raw_data → tidy → write.

from __future__ import annotations

import json
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from builders.db_details_tab import OPTION_PRIORITY as VM_OPTION_PRIORITY, OPTION_SYMBOL as VM_OPTION_SYMBOL


# ── helper ────────────────────────────────────────────────────────────────────
def _extract_vm_fields(r: pd.Series) -> pd.Series:
    """
    Parse the embedded raw_data JSON string (if any) and surface commonly
    requested fields so that analysts can view them without inspecting the
    blob.  Existing non‑blank values in *r* take precedence over JSON values.
    """
    wanted: List[str] = [
        "operating_system_release",
        "cpu_speed",
        "cpu_threads",
        "siblings",
        "hyper_threading",
        "device_model",
        "device_manufacturer",
        "lscpu_total_threads",
        "lscpu_cores_per_socket",
        "lscpu_threads_per_core",
        "lscpu_hypervisor",
    ]

    for k in wanted:
        if k not in r:
            r[k] = ""

    raw = r.get("raw_data", "")
    if not raw:
        return r

    if isinstance(raw, dict):
        j = raw
    else:
        try:
            j = json.loads(raw)
        except Exception:
            import ast
            try:
                j = ast.literal_eval(raw)
            except Exception:
                j = {}

    if isinstance(j, dict):
        for k in wanted:
            if not r.get(k):
                r[k] = j.get(k, "")

    return r


# ── core builder ─────────────────────────────────────────────────────────────
def build_virtual_devices_df(
    virtual_devices_json: str | Path,
    allowed_vms: list[str] | None = None,
    df_db_details: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """
    Load *Virtual Devices.json*, normalise column names, unpack raw_data, and
    return a tidy DataFrame ready for the Excel tab.
    """

    df = pd.read_json(virtual_devices_json)

    # Filter to only allowed VMs if provided
    if allowed_vms is not None:
        df = df[df["device_name"].isin(allowed_vms)]

    # 1. normalise key names and sizing fields
    df = df.rename(
        columns={
            "device_name": "virtual_device",
            "physical_host": "physical_device",
            "total_number_of_processors": "number_of_virtual_processors",
            "total_number_of_cores": "number_of_virtual_cores",
            "total_number_of_threads": "number_of_virtual_threads",
        }
    )

    # ── roll up option symbols from Database Details (wide format) ────────────
    if df_db_details is not None:
        # Ensure df_db_details has a virtual_device column matching device_name
        df_db_details["virtual_device"] = df_db_details["device_name"]
        # Identify option columns in the DB details (excluding known core fields)
        core_db_cols = {
            "device_name", "dbid", "database_name", "product_edition",
            "product_version", "full_version", "instance_status",
            "goldengate_enabled", "source", "rac_hosts", "rac_instances",
            "rac_members_count", "virtual_device"
        }
        option_cols_db = [
            c for c in df_db_details.columns
            if c not in core_db_cols
        ]
        # Group by virtual_device and pick highest-priority symbol per option
        df_opts_grp = df_db_details.groupby("virtual_device")[option_cols_db].agg(
            lambda syms: max(syms, key=lambda sym: VM_OPTION_PRIORITY.get(sym, 0))
            if len(syms.dropna()) > 0 else ""
        ).reset_index()
        # Merge symbols into main DF
        df = df.merge(df_opts_grp, on="virtual_device", how="left")

        # Debug: show which option columns were rolled up
        print("DEBUG: option_cols_db =", option_cols_db)
        print("DEBUG: columns after merge =", df.columns.tolist())
        try:
            print("DEBUG: sample option values:\n", df_opts_grp.head())
        except NameError:
            pass

    # 2. unpack raw_data
    df = df.apply(_extract_vm_fields, axis=1)

    # ── drop redundant or duplicate columns ────────────────────────────────────
    df = df.drop(columns=[
        "cpu_threads", "device_type", "manufacturer", "model",
        "number_of_virtual_cores", "number_of_virtual_threads", "raw_data", "siblings"
    ], errors="ignore")

    # 3. tidy column order: ids / hardware first, everything else later
    id_block = [
        "physical_device",
        "virtual_device",
        "virtualization_type",
        "capped",
        "device_model",
        "device_manufacturer",
        "lscpu_hypervisor",
        "hyper_threading",
        "operating_system_type",
        "operating_system_release",
        "cpu_model",
        "cpu_speed",
        "lscpu_cores_per_socket",
        "lscpu_threads_per_core",
        "lscpu_total_threads",
        "number_of_virtual_processors",
        "oracle_core_factor",
    ]
    id_block = [c for c in id_block if c in df.columns]
    option_block = sorted([c for c in df.columns if c not in id_block])

    df = df[id_block + option_block].fillna("")
    return df


# ── sheet writer ─────────────────────────────────────────────────────────────
def write_virtual_devices_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    sheet_name: str = "Virtual Devices",
) -> None:
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    ws.freeze_panes = ws["A2"]


# ── convenience front‑door ───────────────────────────────────────────────────
def build(
    virtual_devices_json: str | Path,
    allowed_vms: list[str] | None = None,
    df_db_details: pd.DataFrame | None = None,
    workbook: Workbook | None = None,
) -> tuple[Workbook, pd.DataFrame]:
    """
    Convenience wrapper: returns (*workbook*, *df_virtual_devices*).  Creates a
    new Workbook if one isn’t supplied.
    """
    df_vms = build_virtual_devices_df(
        virtual_devices_json,
        allowed_vms=allowed_vms,
        df_db_details=df_db_details
    )
    wb = workbook or Workbook()
    write_virtual_devices_sheet(wb, df_vms)
    return wb, df_vms